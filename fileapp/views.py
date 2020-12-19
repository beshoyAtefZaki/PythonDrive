from django.shortcuts import render ,redirect
from django.http import JsonResponse ,HttpResponse
from .models import FilesData
from django.contrib.auth.models import User
from django.conf import settings
from django.db.models import Q
import os
import glob
import itertools
import openpyxl
# Create your views here.
accepted_type = ["pdf" , 'xls' , 'xlsx' , 'xlt']

def validate(request ):
	file_name = request.GET.get("filename")
	data= { "message" :validate_file(file_name)}
	return JsonResponse(data)

def validate_file(file_name):
	file_type = file_name.split('.')[-1]
	if file_type.lower() not in accepted_type :
		return  "invalid"
	else:
		return  "valid"
def create_file(request):
		file_name = request.FILES.get('formFile')
		validation = validate_file(file_name.name)
		if validation ==  "invalid":
			return HttpResponse("<h1> Invalid File Type </h1>")
		title = request.POST.get('tiltle')
		content = request.POST.get("content")
		new_file = FilesData(   user     = request.user ,
								title    = request.POST.get('tiltle') ,
								content  =request.POST.get("content"),
								file_name= request.FILES.get('formFile'))
		new_file.save()
		return redirect("/profile/%d"%request.user.id)
		data = {"data":[
						{   "user"      : str(new_file.user) ,
						    "title"     : new_file.title ,
						    "content"   : new_file.content,
						    "file_name" : new_file.file_name.url
						}
						]
						}
		return JsonResponse(data)


def user_files(request, idx):
	user = User.objects.get(id=idx)
	files = FilesData.objects.filter(user=user)
	data = create_serializer(files)

	return JsonResponse(data)

def create_serializer(files):
	
	data = {"data":[{   "user"      : str(i.user) ,
						"title"     : i.title ,
						"content"   : i.content,
						"file_name" : i.file_name.url
						}
						for i in files ]
						}
	
	return data



def get_media_root_path():
	root = settings.MEDIA_ROOT 
	files_path = '/documents/'
	return(str(root) +str(files_path))


def serch_in_files(request):
	pattern = request.GET.get("search") 
	
	if pattern :
		query_all = Q(title__icontains=pattern) | Q(content__icontains=pattern) 
		data = {"data":[obj for obj in FilesData.objects.filter(query_all).distinct()]}
		a = get_pdf_files(pattern)
		b = get_excel_files(pattern)
		all_files = a+b
		# if a.get('data') :
		if len(all_files) > 0 :
			for i in all_files:
				try :
					# query_all = Q(title__icontains=pattern) | Q(content__icontains=pattern) | Q(file_name__icontains= i)
					obj = FilesData.objects.get(file_name__icontains= i)
					data['data'].append(obj)
				except:
					pass
		# else :
	 
		# for i in FilesData.objects.filter(query_all):
	
		return render(request , "home.html" , data)

		# all_content = FilesData.objects.filter()
		# return render(request , "home.html" ,{ "data":data})
	else :

		data={ "data" : FilesData.objects.all()}

	return render(request , "home.html" , data)
def get_pdf_files(pattern):
	import PyPDF2
	import re
	os.chdir(get_media_root_path())
	pdf_files = glob.glob('*.pdf')
	data = []
	for file_name in pdf_files:
		obj = PyPDF2.PdfFileReader(file_name)
		numPages = obj.getNumPages()
		for i in range(0, numPages):
			pageObj = obj.getPage(i)
			text = pageObj.extractText()
		for match in re.finditer(pattern, text):
			files = {'file':str(file_name) , "Page no" : (str(i))}
			data.append(file_name)			 				
	return data


def get_excel_files(pattern):
	os.chdir(get_media_root_path())
	excel_files = [glob.glob('*.%s'%i) for i in accepted_type[1::]]
	excel_list =list(itertools.chain(*excel_files))
	data = []
	for file_name in excel_list :
		match =serach_in_exel_files(file_name,pattern)
		if match :
			data.append(match)
		else:pass
	return data









def serach_in_exel_files(file_name , word):
	os.chdir(get_media_root_path())
	wb = openpyxl.load_workbook(file_name)
	all_sheets = wb.get_sheet_names()
	for sheet in all_sheets :
		local_sheet = wb.get_sheet_by_name(sheet)
		row = local_sheet.max_row
		col = local_sheet.max_column
		for row_number in range(1, row+1):
			for col_number in range(1,col+1):
				if str(word) in str(local_sheet.cell(row=row_number, column=col_number).value) :
					return (file_name)
				else:
					pass
	return False