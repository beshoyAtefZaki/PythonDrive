#create excel search method 





import openpyxl




word = "0"

file_name = "Itemtemp.xlsx"
wb = openpyxl.load_workbook(file_name)
all_sheets = wb.get_sheet_names()

for sheet in all_sheets :
	local_sheet = wb.get_sheet_by_name(sheet)
	row = local_sheet.max_row
	col = local_sheet.max_column
	for row_number in range(1, row+1):
		for col_number in range(1,col+1):
			if str(word) in str(local_sheet.cell(row=row_number, column=col_number).value) :
				return (file_name)
				print("Match + row = %s | col = %s"%(row_number , col_number))
			else:
				pass
return False



